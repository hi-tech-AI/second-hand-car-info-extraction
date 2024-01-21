from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from time import sleep
from openpyxl import Workbook
import json

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

def wait_url(driver : webdriver.Chrome, url : str):
    while True:
        cur_url = driver.current_url
        if cur_url == url:
            break
        sleep(1)

service = Service(executable_path = "C:\chromedriver-win64\chromedriver.exe")
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9030")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.6099.225 Safari/537.3")
driver = webdriver.Chrome(options = options, service = service)

wb = Workbook()
sheet = wb.active
# item = ['Marca y modelo del vehículo', 'año', 'kilómetros', 'ubicación', 'tipo de combustible', 'precio de venta', 'vendedor nombre', 'vendedor ubicación', 'vendedor teléfono']
# for i in range(1, 10):
#     sheet.cell(row = 1, column = i).value = item[i-1]

# output = []
start_row = 1
# for id in range(106, 6107):
#     print(f'Pagination number : {id}')
#     driver.get(f'https://www.autocasion.com/coches-ocasion?page={id}')
company_number = 1
# driver.get('https://www.autocasion.com/coches-ocasion')
while True:
    links = Find_Elements(driver, By.TAG_NAME, 'article')
    count = len(links)
    for index in range(count):
        print(f'Company number --> {company_number}')
        print('\n')
        links = Find_Elements(driver, By.TAG_NAME, 'article')
        link = links[index]
        # car_link = link.find_element(By.TAG_NAME, 'a').get_attribute('href')
        # output.append({'link' : car_link})

        car_brand = link.find_element(By.TAG_NAME, 'h2').text
        print(f'car_brand : {car_brand}')
        
        detail_items = link.find_elements(By.TAG_NAME, 'li')
        fuel = detail_items[0].text
        print(f'fuel : {fuel}')
        year = detail_items[1].text
        print(f'year : {year}')
        kilometer = detail_items[2].text.replace('.', '').split(' ')
        print(f'kilometer : {kilometer[0]}')
        location = link.find_element(By.CLASS_NAME, 'provincia').text
        print(f'location : {location}')
        price = link.find_element(By.CLASS_NAME, 'precio ').find_element(By.TAG_NAME, 'span').text.replace('.', '').split(' ')
        print(f'price : {price[0]}')

        get_car = link.find_element(By.TAG_NAME, 'h2')
        driver.execute_script('arguments[0].click();', get_car)
        seller_items = Find_Element(driver, By.CLASS_NAME, 'datos-concesionario').find_elements(By.TAG_NAME, 'p')
        seller_name = seller_items[1].text
        print(f'seller name : {seller_name}')
        try:
            seller_address = driver.find_element(By.CLASS_NAME, 'direccion').text.split('\n')
            print(f'seller address : {" ".join(seller_address[:2])}')
            sheet.cell(row = start_row, column = 8).value = " ".join(seller_address[:2])
        except:
            seller_address = seller_items[2].text
            print(f'seller address : {seller_address}')
            sheet.cell(row = start_row, column = 8).value = seller_address
        try:
            seller_phone = driver.find_element(By.CLASS_NAME, 'datos-concesionario').find_element(By.CLASS_NAME, 'btn-blue-empty').get_attribute('data-phone').replace(' ', '')
            print(f'seller phone : {seller_phone}')
        except:
            pass

        sheet.cell(row = start_row, column = 1).value = car_brand
        sheet.cell(row = start_row, column = 2).value = year
        sheet.cell(row = start_row, column = 3).value = kilometer[0]
        sheet.cell(row = start_row, column = 4).value = location
        sheet.cell(row = start_row, column = 5).value = fuel
        sheet.cell(row = start_row, column = 6).value = price[0]
        sheet.cell(row = start_row, column = 7).value = seller_name
        sheet.cell(row = start_row, column = 9).value = seller_phone
        
        wb.save('output.xlsx')
        start_row += 1
        
        driver.back()
        sleep(1)
        print('\n')
        company_number += 1
    pagination = Find_Element(driver, By.CLASS_NAME, 'paginacion').find_element(By.TAG_NAME, 'ul').find_elements(By.TAG_NAME, 'li')
    next_btn = pagination[2].find_element(By.TAG_NAME, 'a')
    driver.execute_script('arguments[0].click();', next_btn)

    # with open('output.json', 'w') as file:
        # json.dump(output, file)

    # driver.delete_all_cookies()